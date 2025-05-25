from fastapi import APIRouter, Depends, Request, Query, HTTPException, Path, Body
from fastapi.responses import HTMLResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
import crud, schemas, re, models
from database import get_db
import pandas as pd
from fastapi.responses import FileResponse
import os
from datetime import datetime
from urllib.parse import unquote
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font

router = APIRouter()
templates = Jinja2Templates(directory="templates")

# Endpoint do dodawania polisy
@router.post("/dodaj_polise/")
def dodaj_polise(polisa: schemas.PolisaCreate, db: Session = Depends(get_db)):
    print("Otrzymano polise:", polisa)
    return crud.dodaj_polise(db, polisa)

# Endpoint do pobierania polis
@router.get("/polisy", response_model=schemas.PolisaResponse)
def pobierz_polisy(numer_polisy: str = Query(..., description="Numer polisy zakodowany w URL"), db: Session = Depends(get_db)):
    print(f"Zapytanie o polisę z numerem: {numer_polisy}")

    polisa = db.query(models.Polisa).filter(models.Polisa.numer_ubezpieczenia == numer_polisy.strip()).first()
    if not polisa:
        print("Nie znaleziono polisy o podanym numerze.")
        raise HTTPException(status_code=404, detail="Nie znaleziono polisy o podanym numerze.")

    ubezpieczony = db.query(models.Ubezpieczony).filter(models.Ubezpieczony.numer_polisy == numer_polisy.strip()).first()
    if ubezpieczony:
        print(f"Znaleziono ubezpieczonego: {ubezpieczony.ubezpieczony}")
        polisa.ubezpieczony = ubezpieczony.ubezpieczony  
    else:
        print(f"Nie znaleziono ubezpieczonego dla numeru polisy: {numer_polisy}")
        polisa.ubezpieczony = "Brak danych"

    return schemas.PolisaResponse.from_orm(polisa)

# Endpoint do dodawania firmy
@router.post("/dodaj_firme/")
def dodaj_firme(firma: schemas.FirmaCreate, db: Session = Depends(get_db)):
    print("Otrzymano firme:", firma)
    return crud.dodaj_firme(db, firma)

# Endpoint do pobierania 10 pierwszych pozycji z tabeli archiwum
@router.get("/pierwsze_pozycje/", response_model=list[schemas.PolisaResponse])
def pobierz_pierwsze_pozycje(db: Session = Depends(get_db)):
    pozycje = crud.pobierz_pierwsze_pozycje(db)
    print("Pobrane pierwsze pozycje:", pozycje)
    return pozycje

# Endpoint do dodawania ubezpieczonego
@router.post("/dodaj_ubezpieczonego/")
def dodaj_ubezpieczonego(ubezpieczony: schemas.UbezpieczonyCreate, db: Session = Depends(get_db)):
    print("Otrzymano ubezpieczonego:", ubezpieczony)
    return crud.dodaj_ubezpieczonego(db, ubezpieczony)

# Endpoint do wyszukiwania polis
@router.get("/wyszukaj/", response_model=list[schemas.PolisaResponse])
def wyszukaj_polisy(
    ubezpieczajacy: str = None,
    ubezpieczony: str = None,
    nip: str = None,
    regon: str = None,
    przedmiot: str = None,
    numer_polisy: str = None,
    data_zawarcia_od: str = None,
    data_zawarcia_do: str = None,
    ochrona_od: str = None,
    ochrona_do: str = None,
    zakonczenie_od: str = None,
    zakonczenie_do: str = None,
    limit: int = Query(10, ge=1),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db)
):
    wyniki = crud.wyszukaj_polisy(db, ubezpieczajacy, ubezpieczony, nip, regon, przedmiot, numer_polisy, data_zawarcia_od, data_zawarcia_do, ochrona_od, ochrona_do, zakonczenie_od, zakonczenie_do, limit, offset)
    print("Pobrane wyniki wyszukiwania:", wyniki)
    return wyniki

@router.post("/generuj_zestawienie/")
def generuj_zestawienie(
    ubezpieczajacy: str = None,
    ubezpieczony: str = None,
    nip: str = None,
    regon: str = None,
    przedmiot: str = None,
    numer_polisy: str = None,
    data_zawarcia_od: str = None,
    data_zawarcia_do: str = None,
    ochrona_od: str = None,
    ochrona_do: str = None,
    zakonczenie_od: str = None,
    zakonczenie_do: str = None,
    db: Session = Depends(get_db)
):
    wyniki = crud.wyszukaj_polisy(
        db,
        ubezpieczajacy=ubezpieczajacy,
        ubezpieczony=ubezpieczony,
        nip=nip,
        regon=regon,
        przedmiot=przedmiot,
        numer_polisy=numer_polisy,
        data_zawarcia_od=data_zawarcia_od,
        data_zawarcia_do=data_zawarcia_do,
        ochrona_od=ochrona_od,
        ochrona_do=ochrona_do,
        zakonczenie_od=zakonczenie_od,
        zakonczenie_do=zakonczenie_do,
        limit=5000, 
        offset=0
    )
    
    # Logowanie wyników
    print("Wyniki wyszukiwania:", wyniki)

    # Konwertuj wyniki do DataFrame
    df = pd.DataFrame([schemas.PolisaResponse.from_orm(w).dict() for w in wyniki])
    
    # Utwórz folder Zestawienia, jeśli nie istnieje
    folder_path = "Zestawienia"
    os.makedirs(folder_path, exist_ok=True)
    
    # Generuj nazwę pliku
    now = datetime.now()
    file_name = f"Zestawienie {now.strftime('%Y-%m-%d %H_%M')}.xlsx"
    file_path = os.path.join(folder_path, file_name)
    
    # Zapisz DataFrame do pliku Excel
    df.to_excel(file_path, index=False)
    
    return {"message": f"Zestawienie zapisane jako {file_name}"}

# Endpoint do sprawdzania polisy
@router.get("/platnosci", response_model=dict)
def sprawdz_polisy(
    numer_polisy: str = Query(..., description="Numer polisy zakodowany w URL"),
    db: Session = Depends(get_db)
):
    try:
        print(f"Otrzymany numer polisy (zakodowany): {numer_polisy}")

        numer_polisy = unquote(numer_polisy) 
        print(f"Zdekodowany numer polisy: {numer_polisy}")

        platnosc = crud.sprawdz_platnosci(db, numer_polisy)
        if not platnosc:
            print("Nie znaleziono płatności dla podanego numeru polisy.")
            return {"exists": False}

        # Walidacja danych płatności
        if not platnosc.platnosci or platnosc.platnosci.strip() == "":
            print("Brak danych płatności dla tej polisy.")
            return {"exists": True, "platnosci": ""}

        platnosci = platnosc.platnosci.split(";")
        poprawne_platnosci = []
        for p in platnosci:
            try:
                data_skladki, skladka, data_zaplacenia, kwota_zaplacenia = p.split(",")
                poprawne_platnosci.append(f"{data_skladki},{skladka},{data_zaplacenia},{kwota_zaplacenia}")
            except ValueError as e:
                print(f"Błąd podczas przetwarzania płatności: {p}, błąd: {e}")
                continue

        platnosc.platnosci = ";".join(poprawne_platnosci)
        return {"exists": True, "platnosci": platnosc.platnosci}
    except Exception as e:
        print(f"Błąd podczas sprawdzania płatności: {e}")
        raise HTTPException(status_code=500, detail="Wystąpił błąd podczas sprawdzania płatności.")
    
# Endpoint do zapisywania płatności
@router.post("/platnosci/", response_model=schemas.PlatnosciResponse)
def zapisz_platnosci(platnosci: schemas.PlatnosciCreate, db: Session = Depends(get_db)):
    print(f"Otrzymane dane: numer_polisy={platnosci.numer_polisy}, platnosci={platnosci.platnosci}, kurtaz={platnosci.kurtaz}")
    try:
        nowa_platnosc = crud.zapisz_platnosci(db, platnosci)
        return nowa_platnosc
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))

################
# ZMIANY TUTAJ #
################

# Endpoint do aktualizacji płatności
@router.put("/polisy")
def aktualizuj_polisy(policja: schemas.PolisaCreate, db: Session = Depends(get_db)):
    print("Otrzymane dane do aktualizacji:", policja.dict())
    print("Numer polisy:", policja.numer_ubezpieczenia)

    if not policja.numer_ubezpieczenia.strip():
        raise HTTPException(status_code=400, detail="Numer polisy nie może być pusty.")

    polisa = db.query(models.Polisa).filter(models.Polisa.numer_ubezpieczenia == policja.numer_ubezpieczenia).first()
    if not polisa:
        raise HTTPException(status_code=404, detail="Nie znaleziono polisy o podanym numerze.")

    try:
        polisa.ubezpieczajacy = policja.ubezpieczajacy
        polisa.przedmiot_ubezpieczenia = policja.przedmiot_ubezpieczenia
        polisa.ochrona_od = policja.ochrona_od
        polisa.ochrona_do = policja.ochrona_do
        polisa.skladka = policja.skladka

        # Dodaj lub zaktualizuj ubezpieczonego
        ubezpieczony = db.query(models.Ubezpieczony).filter(models.Ubezpieczony.numer_polisy == policja.numer_ubezpieczenia).first()
        if ubezpieczony:
            ubezpieczony.ubezpieczony = policja.ubezpieczony
        else:
            nowy_ubezpieczony = models.Ubezpieczony(
                numer_polisy=polisa.numer_ubezpieczenia,
                ubezpieczony=policja.ubezpieczony
            )
            db.add(nowy_ubezpieczony)

        db.commit()
        print("Polisa zaktualizowana pomyślnie.")
        return {"message": "Polisa zaktualizowana pomyślnie"}
    except Exception as e:
        print("Błąd podczas aktualizacji polisy:", e)
        raise HTTPException(status_code=500, detail="Wystąpił błąd podczas aktualizacji polisy.")
    
@router.post("/raport_knf/")
def generuj_raport_knf(
    data_zawarcia_od: str = Body(..., description="Data zawarcia od"),
    data_zawarcia_do: str = Body(..., description="Data zawarcia do"),
    db: Session = Depends(get_db)
):
    data_zawarcia_od_date = datetime.strptime(data_zawarcia_od, "%Y-%m-%d").date()
    data_zawarcia_do_date = datetime.strptime(data_zawarcia_do, "%Y-%m-%d").date()

    polisy = (
        db.query(models.Polisa.numer_ubezpieczenia, models.Polisa.towarzystwo, models.Polisa.skladka)
        .filter(models.Polisa.data_zawarcia >= data_zawarcia_od_date)
        .filter(models.Polisa.data_zawarcia <= data_zawarcia_do_date)
        .all()
    )

    raport = {}
    for numer_polisy, towarzystwo, skladka in polisy:
        if towarzystwo not in raport:
            raport[towarzystwo] = {
                "liczba_umow": 0, 
                "suma_skladek": 0,
                "suma_skladek_kurtaz": 0,
                "suma_kwot_zaplacenia": 0,
                "prowizja_zainkasowana": 0
            }

        raport[towarzystwo]["liczba_umow"] += 1

        raport[towarzystwo]["suma_skladek"] += skladka

        platnosc = db.query(models.Platnosci).filter(models.Platnosci.numer_polisy == numer_polisy).first()
        if platnosc and platnosc.kurtaz:
            raport[towarzystwo]["suma_skladek_kurtaz"] += Decimal(skladka) * platnosc.kurtaz

        if platnosc and platnosc.platnosci:
            platnosci = platnosc.platnosci.split(";")
            for platnosc_entry in platnosci:
                try:
                    planowana_data, skladka_raty, data_zaplacenia, kwota_zaplacenia = platnosc_entry.split(",")
                    if data_zaplacenia.strip():
                        data_zaplacenia_date = datetime.strptime(data_zaplacenia.strip(), "%Y-%m-%d").date()
                        if data_zawarcia_od_date <= data_zaplacenia_date <= data_zawarcia_do_date:
                            if kwota_zaplacenia.strip():
                                kwota_zaplacenia_decimal = Decimal(kwota_zaplacenia.strip())
                                raport[towarzystwo]["suma_kwot_zaplacenia"] += kwota_zaplacenia_decimal
                                raport[towarzystwo]["prowizja_zainkasowana"] += kwota_zaplacenia_decimal * platnosc.kurtaz
                except ValueError as e:
                    print(f"Błąd przetwarzania płatności: {platnosc_entry}, błąd: {e}")
                    continue

    df = pd.DataFrame(
        [
            {
                "Lp.": idx + 1,
                "Towarzystwo": towarzystwo,
                "Liczba Umów": dane["liczba_umow"],
                "Suma Składek": dane["suma_skladek"],
                "Suma Składek * Kurtaż": dane["suma_skladek_kurtaz"],
                "Suma Kwot Zapłacenia": dane["suma_kwot_zaplacenia"],
                "Prowizja Zainkasowana": dane["prowizja_zainkasowana"]
            }
            for idx, (towarzystwo, dane) in enumerate(raport.items())
        ]
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Raport KNF"

    ws.merge_cells("A1:G1")
    ws["A1"] = f"Raport dla KNF za okres: {data_zawarcia_od} - {data_zawarcia_do}"
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].font = Font(bold=True, size=14)

    ws.merge_cells("C2:D2")
    ws["C2"] = f"Data zawarcia od: {data_zawarcia_od} - {data_zawarcia_do}"
    ws["C2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["C2"].font = Font(bold=True)

    current_date = datetime.now().strftime("%Y-%m-%d")
    ws.merge_cells("E2:G2")
    ws["E2"] = f"Data wygenerowania: {current_date}"
    ws["E2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["E2"].font = Font(bold=True)

    ws.merge_cells("A3:A4")  # Lp.
    ws.merge_cells("B3:B4")  # Towarzystwo
    ws.merge_cells("C3:C4")  # Liczba Umów
    ws.merge_cells("D3:E3")  # Składka plasowana
    ws.merge_cells("F3:G3")  # Prowizja

    ws["A3"] = "Lp."
    ws["B3"] = "Towarzystwo"
    ws["C3"] = "Liczba Umów"
    ws["D3"] = "Składka plasowana"
    ws["F3"] = "Prowizja"
    ws["D4"] = "Przypisana"
    ws["E4"] = "Plasowana"
    ws["F4"] = "Przypisana"
    ws["G4"] = "Zainkasowana"

    for cell in ws["A3:G4"]:
        for c in cell:
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(bold=True)

    ws.column_dimensions["A"].width = 5  # Lp.
    ws.column_dimensions["B"].width = 30  # Towarzystwo
    ws.column_dimensions["C"].width = 15  # Liczba Umów
    ws.column_dimensions["D"].width = 20  # Składka przypisana
    ws.column_dimensions["E"].width = 20  # Składka plasowana
    ws.column_dimensions["F"].width = 20  # Prowizja przypisana
    ws.column_dimensions["G"].width = 20  # Prowizja zainkasowana

    for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=5):
        ws.append(row)

        ws[f"D{row_idx}"].number_format = "#,##0.00 zł"
        ws[f"E{row_idx}"].number_format = "#,##0.00 zł"
        ws[f"F{row_idx}"].number_format = "#,##0.00 zł"
        ws[f"G{row_idx}"].number_format = "#,##0.00 zł"

    last_row = ws.max_row + 1
    ws[f"A{last_row}"] = "Suma"
    ws[f"A{last_row}"].font = Font(bold=True)
    ws[f"A{last_row}"].alignment = Alignment(horizontal="center", vertical="center")

    ws[f"C{last_row}"] = f"=SUM(C5:C{last_row-1})"
    ws[f"C{last_row}"].number_format = "#,##0"

    ws[f"D{last_row}"] = f"=SUM(D5:D{last_row-1})"
    ws[f"D{last_row}"].number_format = "#,##0.00 zł"

    ws[f"E{last_row}"] = f"=SUM(E5:E{last_row-1})"
    ws[f"E{last_row}"].number_format = "#,##0.00 zł"

    ws[f"F{last_row}"] = f"=SUM(F5:F{last_row-1})"
    ws[f"F{last_row}"].number_format = "#,##0.00 zł"

    ws[f"G{last_row}"] = f"=SUM(G5:G{last_row-1})"
    ws[f"G{last_row}"].number_format = "#,##0.00 zł"

    folder_path = "RaportyKNF"
    os.makedirs(folder_path, exist_ok=True)
    now = datetime.now()
    file_name = f"Raport_KNF_{data_zawarcia_od}_do_{data_zawarcia_do}_{now.strftime('%H_%M')}.xlsx"
    file_path = os.path.join(folder_path, file_name)
    wb.save(file_path)

    return {"message": f"Raport wygenerowany: {file_name}", "file_path": file_path}

@router.get("/szukaj_firme/", response_model=schemas.FirmaResponse)
def szukaj_firme(nip: str = Query(..., description="NIP firmy"), db: Session = Depends(get_db)):
    firma = db.query(models.Firma).filter(models.Firma.nip == nip.strip()).first()
    if not firma:
        raise HTTPException(status_code=404, detail="Nie znaleziono firmy o podanym NIP.")
    return firma

@router.get("/szukaj_firme_regon/", response_model=schemas.FirmaResponse)
def szukaj_firme_regon(regon: str = Query(..., description="REGON firmy"), db: Session = Depends(get_db)):
    firma = db.query(models.Firma).filter(models.Firma.regon == regon.strip()).first()
    if not firma:
        raise HTTPException(status_code=404, detail="Nie znaleziono firmy o podanym REGON.")
    return firma

@router.post("/notatki/", response_model=schemas.NotatkaResponse)
def dodaj_notatke(notatka: schemas.NotatkaCreate, db: Session = Depends(get_db)):
    return crud.dodaj_notatke(db, notatka)

@router.get("/notatki/{numer_polisy}", response_model=list[schemas.NotatkaResponse])
def pobierz_notatki(numer_polisy: str, db: Session = Depends(get_db)):
    return crud.pobierz_notatki(db, numer_polisy)

################
# ZMIANY TUTAJ #
################

@router.put("/platnosci")
def aktualizuj_platnosci(
    numer_polisy: str = Query(..., description="Numer polisy"),
    nowe_platnosci: dict = Body(...),
    db: Session = Depends(get_db)
):
    try:
        platnosc = db.query(models.Platnosci).filter(models.Platnosci.numer_polisy == numer_polisy).first()
        if not platnosc:
            raise HTTPException(status_code=404, detail="Nie znaleziono płatności dla tej polisy.")

        # Zbuduj nowy tekst do pola platnosci na podstawie przesłanych danych
        platnosci_lista = []
        for i in range(len(nowe_platnosci)):
            p = nowe_platnosci[str(i)]
            platnosci_lista.append(
                f"{p['dataPlatnosci']},{p['skladka']},{p['dataZaplacenia']},{p['kwotaZaplacenia']}"
            )
        platnosc.platnosci = ";".join(platnosci_lista)

        db.commit()
        return {"message": "Płatności zaktualizowane pomyślnie."}
    except Exception as e:
        print("Błąd podczas aktualizacji płatności:", e)
        raise HTTPException(status_code=500, detail="Wystąpił błąd podczas aktualizacji płatności.")